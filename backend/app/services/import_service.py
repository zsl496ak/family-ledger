import csv
import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..models import Account, Category, Transaction


def import_transactions_file(db: Session, content: bytes, filename: str, family_id: int, creator_id: int) -> dict:
    rows = _parse_file(content, filename)
    accounts = {a.name: a.id for a in db.query(Account).filter(Account.family_id == family_id).all()}
    categories = {c.name: c.id for c in db.query(Category).filter(Category.family_id == family_id).all()}

    # Get existing transactions for dedup
    existing = set()
    for t in db.query(Transaction).filter(Transaction.family_id == family_id).all():
        existing.add((str(t.transaction_date), str(t.amount), t.note or ""))

    success = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        try:
            row_date = _parse_date(row.get("日期", ""))
            row_type_raw = row.get("类型", "").strip()
            row_type = {"收入": "income", "支出": "expense", "转账": "transfer",
                        "income": "income", "expense": "expense", "transfer": "transfer"}.get(row_type_raw)
            if not row_type:
                errors.append({"row": i, "error": f"无效的类型: {row_type_raw}"})
                continue

            category_name = row.get("分类", "").strip()
            account_name = row.get("账户", "").strip()
            amount_str = row.get("金额", "").strip()
            note = row.get("备注", "").strip()

            if not account_name:
                errors.append({"row": i, "error": "账户不能为空"})
                continue
            if not amount_str:
                errors.append({"row": i, "error": "金额不能为空"})
                continue

            try:
                amount = Decimal(amount_str)
            except InvalidOperation:
                errors.append({"row": i, "error": f"无效的金额: {amount_str}"})
                continue

            if amount <= 0:
                errors.append({"row": i, "error": "金额必须大于0"})
                continue

            account_id = accounts.get(account_name)
            if not account_id:
                errors.append({"row": i, "error": f"账户不存在: {account_name}"})
                continue

            category_id = categories.get(category_name) if category_name else None

            # Dedup check
            dedup_key = (str(row_date), str(amount), note)
            if dedup_key in existing:
                skipped += 1
                continue

            txn = Transaction(
                family_id=family_id, account_id=account_id,
                category_id=category_id, transaction_type=row_type,
                amount=amount, note=note or None,
                transaction_date=row_date, creator_id=creator_id,
            )
            db.add(txn)
            existing.add(dedup_key)
            success += 1

        except Exception as e:
            errors.append({"row": i, "error": str(e)})

    db.commit()
    return {"success": success, "skipped": skipped, "errors": errors, "total": len(rows)}


def _parse_file(content: bytes, filename: str) -> list[dict]:
    if filename.endswith(".xlsx"):
        return _parse_xlsx(content)
    elif filename.endswith(".csv"):
        return _parse_csv(content)
    else:
        # Try xlsx first, then csv
        try:
            return _parse_xlsx(content)
        except Exception:
            return _parse_csv(content)


def _parse_xlsx(content: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return []

    headers = [str(h).strip().rstrip("*").strip() if h else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        if not any(row):
            continue
        d = {}
        for j, val in enumerate(row):
            if j < len(headers) and headers[j]:
                if isinstance(val, datetime):
                    d[headers[j]] = val.strftime("%Y-%m-%d")
                elif isinstance(val, date):
                    d[headers[j]] = val.isoformat()
                elif val is not None:
                    d[headers[j]] = str(val)
                else:
                    d[headers[j]] = ""
        result.append(d)
    wb.close()
    return result


def _parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        rows.append({k.strip().rstrip("*").strip() if k else "": v for k, v in row.items()})
    return rows


def _parse_date(s: str) -> date:
    s = s.strip()
    if not s:
        raise ValueError("日期不能为空")
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日", "%Y.%m.%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        num = float(s)
        if num > 30000 and num < 100000:
            return date(1899, 12, 30) + __import__("datetime").timedelta(days=int(num))
    except ValueError:
        pass
    raise ValueError(f"无法解析日期: {s}")
