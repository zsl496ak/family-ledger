import io
import csv as csv_module
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..database import get_db
from ..models import User, Transaction as TransactionModel
from ..schemas.transaction import TransactionCreate, TransactionUpdate, TransactionOut, TransactionFilter, TransactionSummary
from ..schemas.common import PaginatedResponse
from ..services import transaction_service
from .deps import get_current_user

router = APIRouter(prefix="/transactions", tags=["transactions"])


@router.get("")
def list_transactions(
    date_from: date = None,
    date_to: date = None,
    transaction_type: str = None,
    category_id: int = None,
    account_id: int = None,
    search: str = None,
    page: int = 1,
    page_size: int = 20,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    filters = TransactionFilter(
        date_from=date_from, date_to=date_to, transaction_type=transaction_type,
        category_id=category_id, account_id=account_id, search=search,
        page=page, page_size=page_size,
    )
    items, total = transaction_service.get_transactions(db, current_user.family_id, filters)
    return PaginatedResponse(
        items=[TransactionOut.model_validate(item) for item in items],
        total=total, page=page, page_size=page_size,
        total_pages=(total + page_size - 1) // page_size,
    )


@router.post("", response_model=TransactionOut)
def create_transaction(req: TransactionCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return transaction_service.create_transaction(db, current_user.family_id, current_user.id, req.model_dump())


@router.get("/summary", response_model=TransactionSummary)
def get_summary(
    date_from: date = None, date_to: date = None,
    current_user: User = Depends(get_current_user), db: Session = Depends(get_db),
):
    return transaction_service.get_summary(db, current_user.family_id, date_from, date_to)


@router.get("/{transaction_id}", response_model=TransactionOut)
def get_transaction(transaction_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    txn = transaction_service.get_transaction(db, transaction_id, current_user.family_id)
    if not txn:
        raise HTTPException(status_code=404, detail="记录不存在")
    return TransactionOut.model_validate(txn)


@router.put("/{transaction_id}", response_model=TransactionOut)
def update_transaction(transaction_id: int, req: TransactionUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    txn = transaction_service.get_transaction(db, transaction_id, current_user.family_id)
    if not txn:
        raise HTTPException(status_code=404, detail="记录不存在")
    return transaction_service.update_transaction(db, txn, req.model_dump(exclude_unset=True))


@router.delete("/{transaction_id}")
def delete_transaction(transaction_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    txn = transaction_service.get_transaction(db, transaction_id, current_user.family_id)
    if not txn:
        raise HTTPException(status_code=404, detail="记录不存在")
    transaction_service.delete_transaction(db, txn)
    return {"message": "删除成功"}


@router.post("/export")
def export_transactions(
    date_from: date = None, date_to: date = None,
    transaction_type: str = None, format: str = "xlsx",
    current_user: User = Depends(get_current_user), db: Session = Depends(get_db),
):
    filters = TransactionFilter(date_from=date_from, date_to=date_to, transaction_type=transaction_type, page_size=10000)
    items, _ = transaction_service.get_transactions(db, current_user.family_id, filters)

    if format == "csv":
        output = io.StringIO()
        writer = csv_module.writer(output)
        writer.writerow(["日期", "类型", "分类", "账户", "金额", "备注", "记账人"])
        for item in items:
            writer.writerow([
                item.transaction_date, item.transaction_type,
                item.category_name or "", item.account_name or "",
                item.amount, item.note or "", item.creator_name or "",
            ])
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=transactions.csv"},
        )

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "交易记录"
    ws.append(["日期", "类型", "分类", "账户", "金额", "备注", "记账人"])
    for item in items:
        type_map = {"income": "收入", "expense": "支出", "transfer": "转账"}
        ws.append([
            str(item.transaction_date), type_map.get(item.transaction_type, item.transaction_type),
            item.category_name or "", item.account_name or "",
            float(item.amount), item.note or "", item.creator_name or "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=transactions.xlsx"},
    )


@router.post("/import")
async def import_transactions(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    from ..services.import_service import import_transactions_file
    content = await file.read()
    result = import_transactions_file(db, content, file.filename, current_user.family_id, current_user.id)
    return result


@router.get("/import/template")
def download_template(current_user: User = Depends(get_current_user)):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "交易导入模板"
    ws.append(["日期*", "类型*", "分类", "账户*", "金额*", "备注"])
    ws.append(["2026-01-15", "expense", "餐饮", "现金", "35.50", "午餐"])
    ws.append(["2026-01-15", "income", "工资", "银行卡", "10000", "1月工资"])
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20

    ws_notes = wb.create_sheet("说明")
    ws_notes.append(["字段", "说明", "示例"])
    ws_notes.append(["日期", "交易日期，格式 YYYY-MM-DD", "2026-01-15"])
    ws_notes.append(["类型", "收入/支出/转账", "expense"])
    ws_notes.append(["分类", "分类名称（需已存在）", "餐饮"])
    ws_notes.append(["账户", "账户名称（需已存在）", "现金"])
    ws_notes.append(["金额", "金额，正数", "35.50"])
    ws_notes.append(["备注", "备注信息", "午餐"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=import_template.xlsx"},
    )
